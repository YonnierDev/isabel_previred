import logging
import os
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_ACCENT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


def _ajustar_columnas(ws, df: pd.DataFrame, extra: int = 4):
    for i, col in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col)),
            df[col].astype(str).str.len().max() if not df.empty else 0,
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + extra, 50)


def _estilizar_encabezados(ws, num_cols: int):
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _estilizar_filas_pares(ws, num_rows: int, num_cols: int):
    for row in range(2, num_rows + 2):
        if row % 2 == 0:
            for col in range(1, num_cols + 1):
                ws.cell(row=row, column=col).fill = _ACCENT_FILL


def _escribir_df(ws, df: pd.DataFrame):
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    _estilizar_encabezados(ws, len(df.columns))
    _estilizar_filas_pares(ws, len(df), len(df.columns))
    _ajustar_columnas(ws, df)


class ReportGenerator:

    def __init__(self, config: dict):
        self.config = config

    def generar_excel(self, datos_procesados: pd.DataFrame, datos_agrupados: dict) -> None:
        import pathlib
        archivo_salida = str(pathlib.Path(self.config["archivo_salida"]).with_suffix(".xlsx"))
        os.makedirs(os.path.dirname(archivo_salida), exist_ok=True)

        df = datos_agrupados.get("por_semana_asesor", pd.DataFrame())

        with pd.ExcelWriter(archivo_salida, engine="openpyxl") as writer:
            ws = writer.book.create_sheet("Informe")
            _escribir_df(ws, df)

        logger.info(f"Reporte generado: {archivo_salida}")

    def _hoja_resumen_asesores(self, writer, df: pd.DataFrame):
        ws = writer.book.create_sheet("Resumen Asesores")
        titulo = f"Resumen por Asesor — {self.config.get('mes', '')}"
        ws.cell(row=1, column=1, value=titulo).font = Font(bold=True, size=13)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns) or 3)
        ws.append([])
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=3, column=col_idx, value=col_name)
        for row_idx, row in enumerate(df.itertuples(index=False), start=4):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        for cell in ws[3]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        _ajustar_columnas(ws, df)

    def _hoja_resumen_semanas(self, writer, df: pd.DataFrame):
        ws = writer.book.create_sheet("Resumen Semanas")
        titulo = f"Resumen por Semana — {self.config.get('mes', '')}"
        ws.cell(row=1, column=1, value=titulo).font = Font(bold=True, size=13)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns) or 3)
        ws.append([])
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=3, column=col_idx, value=col_name)
        for row_idx, row in enumerate(df.itertuples(index=False), start=4):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        for cell in ws[3]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        _ajustar_columnas(ws, df)

    def _hoja_detalles_semana(self, writer, df: pd.DataFrame):
        ws = writer.book.create_sheet("Detalles por Semana")
        _escribir_df(ws, df)

    def _hoja_datos_crudos(self, writer, df: pd.DataFrame):
        cols_mostrar = [c for c in ["Fecha", "Nombre_Asesor", "Duración_Segundos", "Contestada", "Mayor_30_Segundos"] if c in df.columns]
        df_export = df[cols_mostrar].copy() if cols_mostrar else df.copy()
        if "Fecha" in df_export.columns:
            df_export["Fecha"] = df_export["Fecha"].dt.strftime("%Y-%m-%d %H:%M:%S").fillna("")
        ws = writer.book.create_sheet("Datos Crudos")
        _escribir_df(ws, df_export)

    def _hoja_totales(self, writer, df: pd.DataFrame):
        ws = writer.book.create_sheet("Totales")

        mes = self.config.get("mes", "")
        total_contestadas = int(df["Contestada"].sum()) if "Contestada" in df.columns else 0
        total_30s = int(df["Mayor_30_Segundos"].sum()) if "Mayor_30_Segundos" in df.columns else 0
        pct = round(total_30s / total_contestadas * 100, 1) if total_contestadas else 0.0

        kpis = [
            ("Mes de análisis", mes),
            ("Fecha de generación", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("", ""),
            ("Total llamadas contestadas", total_contestadas),
            (f"Llamadas >= {self.config.get('duracion_minima_segundos', 30)} segundos", total_30s),
            ("% Llamadas efectivas", f"{pct}%"),
        ]

        ws.cell(row=1, column=1, value="RESUMEN EJECUTIVO").font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).fill = _HEADER_FILL
        ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFF")
        ws.merge_cells("A1:B1")

        for idx, (label, valor) in enumerate(kpis, start=3):
            ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=idx, column=2, value=valor)

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 28
